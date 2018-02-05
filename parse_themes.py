from openpyxl import load_workbook
from openpyxl import Workbook

# number of interviews
ints = 18

theme_dict = {1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 5, 9: 6, 10: 7, 11: 7, 12: 6, 13: 8, 14: 5, 15: 5, 16: 5, 17: 5, 18: 9}
theme_dict_codes = {1: [], 2: [], 3: [], 4: [], 5: [], 6: [], 7: [], 8: [], 9: []}
theme_dict_answers = {1: [], 2: [], 3: [], 4: [], 5: [], 6: [], 7: [], 8: [], 9: []}
theme_dict_codes_and_answers = {1: [], 2: [], 3: [], 4: [], 5: [], 6: [], 7: [], 8: [], 9: []}
code_dict_for_each_q = {1: {}, 2: {}, 3: {}, 4: {}, 5: {}, 6: {}, 9: {}, 10: {}, 11: {}, 12: {}, 13: {}, 14: {}, 15: {},
                        16: {},
                        17: {}, 18: {}}
answers_dict_for_each_q = {1: {}, 2: {}, 3: {}, 4: {}, 5: {}, 6: {}, 9: {}, 10: {}, 11: {}, 12: {}, 13: {}, 14: {},
                           15: {}, 16: {},
                           17: {}, 18: {}}
all_codes_for_one_q = {1: [], 2: [], 3: [], 4: [], 5: [], 6: [], 9: [], 10: [], 11: [], 12: [], 13: [], 14: [],
                       15: [], 16: [],
                       17: [], 18: []}
all_answers_for_one_q = {1: [], 2: [], 3: [], 4: [], 5: [], 6: [], 9: [], 10: [], 11: [], 12: [], 13: [], 14: [],
                         15: [], 16: [],
                         17: [], 18: []}
translation_table = dict.fromkeys(map(ord, '.!?()'), None)
full_list = []


def populate_variables(ws, ):
    for q_row in range(8, 25):
        q_num = q_row - 7
        for ans_col in range(6, (ints * 2) + 6, 2):
            each_guy_codes = ws.cell(row=q_row, column=ans_col).value
            each_guy_ans = ws.cell(row=q_row, column=ans_col - 1).value
            each_guy_num = int((ans_col / 2) - 2)
            try:
                each_guy_ans = each_guy_ans.translate(translation_table)
                each_guy_codes = each_guy_codes.translate(translation_table)
                lines = each_guy_codes.splitlines()
                codes = [str(str(each_guy_num) + ' ' + line) for line in lines]
                code_dict_for_each_q[q_num][each_guy_num] = codes
                answers_dict_for_each_q[q_num][each_guy_num] = each_guy_ans
                all_codes_for_one_q[q_num].extend(codes)
                all_answers_for_one_q[q_num].extend(each_guy_ans)
                theme_dict_codes[theme_dict[q_num]].extend(codes)
                theme_dict_answers[theme_dict[q_num]].append(each_guy_ans)
                full_list.extend(lines)
            except:
                pass

    for i in range(1, 9):
        theme_dict_codes_and_answers[i] = theme_dict_answers[i] + theme_dict_codes[i]
        # with open('./' + str(i) + '.txt', 'w',encoding='utf8') as f:
        #     f.writelines(theme_dict_answers[i])


def write_each_theme_to_a_col(dic, worksheet, workbook, result_file_address):
    for i in range(1, 9):
        counter = 2
        for val in dic[i]:
            worksheet.cell(row=counter, column=i * 2 - 1).value = 1
            worksheet.cell(row=counter, column=i * 2).value = val
            counter += 1
    workbook.save(result_file_address)


def write_full_list_to_file(full_list, worksheet, workbook, result_file_address):
    for i, l in enumerate(full_list):
        worksheet.cell(row=(i + 2), column=1).value = l
    workbook.save(result_file_address)


def count_freq(theme_dict_codes_and_answers):
    all_words_in_themes_removed = {}
    all_words_in_themes = {}

    for code_num in range(1, 9):
        l = [code.split() for code in theme_dict_codes_and_answers[code_num]]
        all_words_in_themes[code_num] = [item for sublist in l for item in sublist if not item.isdigit()]
        all_words_in_themes_removed[code_num] = [item.replace('های', '').replace('\u200c', '') for sublist in l for item
                                                 in sublist if
                                                 item not in ['در', 'اجتماعی', 'از', 'برای', 'استفاده', 'است', 'باشد',
                                                              'شبکه', 'و', 'به', 'هستند', 'ها', 'با', 'های', 'را', '1',
                                                              '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12',
                                                              '13', '14',
                                                              '15', '16', '17', '18', 'می', 'شود', 'این', 'شده', 'شدند',
                                                              'بودن', 'هم', 'که', 'یا', 'بیشتر', 'یک', 'آن', 'خیلی',
                                                              'ما', 'شد', 'بود', 'اینکه', 'من', 'همه', 'اگر', '.', 'كه',
                                                              'ی', 'مثلا', 'تر', 'تا', 'ای',
                                                              'چه', 'هر', 'مى', 'دارد', 'هر', 'خود', 'اين', 'داشت',
                                                              'حتى',
                                                              'البته', 'بودند', 'رو', 'اما', 'باید', 'دیگر', 'میشود',
                                                              'وقتی',
                                                              '،', 'چون', 'خیر', 'نمی', 'می', 'هاى', '', '',
                                                              'کردن', '10', 'شدن', 'کنند', 'گذاشته', '11', '12',
                                                              'نداشت',
                                                              'ولی', 'دارند', 'شبکه\u200cهای\u200cاجتماعی',
                                                              'می\u200cشد', 'می\u200cشود',
                                                              'صورت', 'هست', 'طریق', 'اعضای', 'داریم', 'شما', 'نه', '',
                                                              '', '', ''
                                                              ]]

        with open('./' + str(code_num) + '.txt', 'w', encoding='utf8') as f:
            f.writelines("\n".join(all_words_in_themes_removed[code_num]))

    import nltk
    w = Workbook()
    for code_num in range(1, 9):
        sheet = w.create_sheet(str(code_num))
        freq2 = nltk.trigrams(all_words_in_themes[code_num])
        ff3 = nltk.FreqDist(freq2)
        res3 = [str(
            item[0][0].replace('\u200c', ' ') + ' ' + item[0][1].replace('\u200c', ' ') + ' ' + item[0][2].replace(
                '\u200c', ' ')) for item in ff3.most_common(50)]
        print(res3)
        for i in range(50):
            sheet.cell(row=i + 1, column=1).value = res3[i]
            sheet.cell(row=i + 1, column=2).value = ff3.most_common(50)[i][1]
        print(ff3.most_common(10))
        freq2 = nltk.bigrams(all_words_in_themes[code_num])
        ff2 = nltk.FreqDist(freq2)
        res2 = [str(item[0][0].replace('\u200c', ' ') + ' ' + item[0][1].replace('\u200c', ' ')) for item in
                ff2.most_common(50)]
        for i in range(50):
            sheet.cell(row=i + 1, column=3).value = res2[i]
            sheet.cell(row=i + 1, column=4).value = ff2.most_common(50)[i][1]
        print(ff2.most_common(10))
        print('top 10 words for theme number : ', code_num)
        freq = nltk.FreqDist(all_words_in_themes_removed[code_num])
        for i in range(10):
            sheet.cell(row=i + 1, column=5).value = freq.most_common(10)[i][0]
            sheet.cell(row=i + 1, column=6).value = freq.most_common(10)[i][1]
        print(freq.most_common(10))
    w.save('balances.xlsx')


if __name__ == "__main__":
    wb = load_workbook('E:\Freax\SelfStuff\Thesis\\beyond\F_18.xlsx')
    ws = wb.active
    wb2 = load_workbook('test3.xlsx')
    ws2 = wb2.active
    populate_variables(ws)
    # write_full_list_to_file(full_list, ws2, wb2, 'test2.xlsx')
    # write_each_theme_to_a_col(theme_dict_codes, ws2, wb2, 'test3.xlsx')
    count_freq(theme_dict_codes_and_answers)
